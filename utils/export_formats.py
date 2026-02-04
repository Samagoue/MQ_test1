"""
Multi-Format Export Utilities

Supports exporting diagrams and data to multiple formats:
- SVG (Scalable Vector Graphics)
- PNG (Portable Network Graphics)
- Excel (XLSX) reports
"""

import subprocess
import shutil
import re
from pathlib import Path
from typing import Dict, List
from datetime import datetime


def _select_layout_engine(dot_file: Path, layout_engine: str = None) -> str:
    """
    Select appropriate GraphViz layout engine.

    Args:
        dot_file: Path to DOT file
        layout_engine: Optional explicit layout engine ('dot', 'sfdp', 'neato', etc.)

    Returns:
        Name of layout engine to use
    """
    if layout_engine:
        return layout_engine

    # Use sfdp for large hierarchical layouts (topology files)
    # Use dot for everything else (better for smaller, structured graphs)
    stem_lower = dot_file.stem.lower()
    if 'topology' in stem_lower or 'hierarchical' in stem_lower:
        return 'sfdp'
    return 'dot'


def export_dot_to_svg(dot_file: Path, svg_file: Path = None, layout_engine: str = None) -> bool:
    """
    Export GraphViz DOT file to SVG format.

    Args:
        dot_file: Path to DOT file
        svg_file: Output SVG file path (optional, defaults to same name as dot_file)
        layout_engine: Optional explicit layout engine ('dot', 'sfdp', etc.)

    Returns:
        True if successful, False otherwise
    """
    if svg_file is None:
        svg_file = dot_file.with_suffix('.svg')

    # Check if dot command exists
    if not shutil.which('dot') and not shutil.which('sfdp'):
        print("⚠ GraphViz not found - cannot generate SVG")
        return False

    try:
        # Select appropriate layout engine
        engine = _select_layout_engine(dot_file, layout_engine)

        subprocess.run(
            [engine, '-Tsvg', str(dot_file), '-o', str(svg_file)],
            check=True,
            capture_output=True,
            text=True
        )

        # Post-process SVG to remove underlines from hyperlinks
        _remove_svg_link_underlines(svg_file)

        print(f"✓ SVG generated: {svg_file}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"✗ SVG generation failed: {e.stderr}")
        return False
    except Exception as e:
        print(f"✗ SVG generation failed: {e}")
        return False


def _remove_svg_link_underlines(svg_file: Path) -> None:
    """
    Post-process SVG file to remove underlines from hyperlinks.

    Injects CSS into the SVG to set text-decoration: none on all links.

    Args:
        svg_file: Path to SVG file to modify
    """
    try:
        content = svg_file.read_text(encoding='utf-8')

        # CSS to remove underlines from links
        css_style = '''<defs>
<style type="text/css">
a { text-decoration: none !important; }
a:hover { text-decoration: none !important; }
a text { text-decoration: none !important; }
</style>
</defs>'''

        # Insert CSS after opening <svg> tag
        if '<defs>' not in content:
            # Insert after the opening <svg ...> tag
            content = re.sub(
                r'(<svg[^>]*>)',
                r'\1\n' + css_style,
                content,
                count=1
            )
            svg_file.write_text(content, encoding='utf-8')
    except Exception as e:
        # Non-fatal - SVG still works, just with underlines
        print(f"  ⚠ Could not remove link underlines: {e}")


def export_dot_to_png(dot_file: Path, png_file: Path = None, dpi: int = 150, layout_engine: str = None) -> bool:
    """
    Export GraphViz DOT file to PNG format.

    Args:
        dot_file: Path to DOT file
        png_file: Output PNG file path (optional, defaults to same name as dot_file)
        dpi: Resolution in dots per inch (default: 150)
        layout_engine: Optional explicit layout engine ('dot', 'sfdp', etc.)

    Returns:
        True if successful, False otherwise
    """
    if png_file is None:
        png_file = dot_file.with_suffix('.png')

    # Check if dot command exists
    if not shutil.which('dot') and not shutil.which('sfdp'):
        print("⚠ GraphViz not found - cannot generate PNG")
        return False

    try:
        # Select appropriate layout engine
        engine = _select_layout_engine(dot_file, layout_engine)

        subprocess.run(
            [engine, '-Tpng', f'-Gdpi={dpi}', str(dot_file), '-o', str(png_file)],
            check=True,
            capture_output=True,
            text=True
        )
        print(f"✓ PNG generated: {png_file}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"✗ PNG generation failed: {e.stderr}")
        return False
    except Exception as e:
        print(f"✗ PNG generation failed: {e}")
        return False


def export_directory_to_formats(directory: Path, formats: List[str] = ['svg', 'png'], dpi: int = 150):
    """
    Export all DOT files in a directory to multiple formats.

    Args:
        directory: Directory containing DOT files
        formats: List of formats to export to ('svg', 'png')
        dpi: Resolution for PNG exports (default: 150)
    """
    if not directory.exists():
        print(f"⚠ Directory not found: {directory}")
        return

    dot_files = list(directory.glob('*.dot'))
    if not dot_files:
        print(f"⚠ No DOT files found in {directory}")
        return

    success_count = {fmt: 0 for fmt in formats}
    total = len(dot_files)

    for dot_file in dot_files:
        for fmt in formats:
            if fmt == 'svg':
                if export_dot_to_svg(dot_file):
                    success_count['svg'] += 1
            elif fmt == 'png':
                if export_dot_to_png(dot_file, dpi=dpi):
                    success_count['png'] += 1

    print(f"\n✓ Export Summary:")
    for fmt in formats:
        print(f"  {fmt.upper()}: {success_count[fmt]}/{total} files")


def generate_excel_inventory(enriched_data: Dict, output_file: Path) -> bool:
    """
    Generate Excel inventory report with multiple sheets.

    Requires openpyxl package: pip install openpyxl

    Args:
        enriched_data: Enriched MQ CMDB data
        output_file: Output Excel file path

    Returns:
        True if successful, False otherwise
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠ openpyxl not installed - Excel export not available")
        print("  Install with: pip install openpyxl")
        return False

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Extract all MQ managers
    all_mqmanagers = []
    all_connections = []
    all_gateways = []

    for org_name, org_data in enriched_data.items():
        if not isinstance(org_data, dict) or '_departments' not in org_data:
            continue

        for dept_name, dept_data in org_data['_departments'].items():
            for biz_ownr, applications in dept_data.items():
                for app_name, mqmgr_dict in applications.items():
                    for mqmgr_name, mqmgr_data in mqmgr_dict.items():
                        # MQ Manager inventory
                        all_mqmanagers.append({
                            'MQ Manager': mqmgr_name,
                            'Organization': mqmgr_data.get('Organization', ''),
                            'Department': mqmgr_data.get('Department', ''),
                            'Biz Owner': mqmgr_data.get('Biz_Ownr', ''),
                            'Application': mqmgr_data.get('Application', ''),
                            'Is Gateway': 'Yes' if mqmgr_data.get('IsGateway', False) else 'No',
                            'Gateway Scope': mqmgr_data.get('GatewayScope', ''),
                            'QLocal Count': mqmgr_data.get('qlocal_count', 0),
                            'QRemote Count': mqmgr_data.get('qremote_count', 0),
                            'QAlias Count': mqmgr_data.get('qalias_count', 0),
                            'Total Queues': mqmgr_data.get('total_count', 0),
                            'Inbound Connections': len(mqmgr_data.get('inbound', [])) + len(mqmgr_data.get('inbound_extra', [])),
                            'Outbound Connections': len(mqmgr_data.get('outbound', [])) + len(mqmgr_data.get('outbound_extra', []))
                        })

                        # Gateway inventory
                        if mqmgr_data.get('IsGateway', False):
                            all_gateways.append({
                                'Gateway Name': mqmgr_name,
                                'Scope': mqmgr_data.get('GatewayScope', ''),
                                'Organization': mqmgr_data.get('Organization', ''),
                                'Department': mqmgr_data.get('Department', ''),
                                'Description': mqmgr_data.get('GatewayDescription', ''),
                                'Total Connections': (len(mqmgr_data.get('inbound', [])) + len(mqmgr_data.get('outbound', [])) +
                                                    len(mqmgr_data.get('inbound_extra', [])) + len(mqmgr_data.get('outbound_extra', [])))
                            })

                        # Connections
                        for target in mqmgr_data.get('outbound', []) + mqmgr_data.get('outbound_extra', []):
                            all_connections.append({
                                'Source': mqmgr_name,
                                'Target': target,
                                'Source Org': mqmgr_data.get('Organization', ''),
                                'Source Dept': mqmgr_data.get('Department', ''),
                                'Source App': mqmgr_data.get('Application', '')
                            })

    # Sheet 1: MQ Manager Inventory
    ws1 = wb.create_sheet("MQ Managers")
    if all_mqmanagers:
        headers = list(all_mqmanagers[0].keys())
        ws1.append(headers)

        # Style header row
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for mqmgr in all_mqmanagers:
            ws1.append(list(mqmgr.values()))

        # Auto-adjust column widths
        for idx, col in enumerate(ws1.columns, 1):
            max_length = 0
            column = get_column_letter(idx)
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (TypeError, AttributeError):
                    # Skip cells that can't be converted to string
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column].width = adjusted_width

    # Sheet 2: Connections
    ws2 = wb.create_sheet("Connections")
    if all_connections:
        headers = list(all_connections[0].keys())
        ws2.append(headers)

        # Style header row
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for conn in all_connections:
            ws2.append(list(conn.values()))

        # Auto-adjust column widths
        for idx, col in enumerate(ws2.columns, 1):
            max_length = 0
            column = get_column_letter(idx)
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (TypeError, AttributeError):
                    # Skip cells that can't be converted to string
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column].width = adjusted_width

    # Sheet 3: Gateway Inventory
    if all_gateways:
        ws3 = wb.create_sheet("Gateways")
        headers = list(all_gateways[0].keys())
        ws3.append(headers)

        # Style header row
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for gw in all_gateways:
            ws3.append(list(gw.values()))

        # Auto-adjust column widths
        for idx, col in enumerate(ws3.columns, 1):
            max_length = 0
            column = get_column_letter(idx)
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (TypeError, AttributeError):
                    # Skip cells that can't be converted to string
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column].width = adjusted_width

    # Sheet 4: Summary Statistics
    ws4 = wb.create_sheet("Summary", 0)  # Insert as first sheet
    ws4['A1'] = 'MQ CMDB Inventory Report'
    ws4['A1'].font = Font(bold=True, size=16)
    ws4['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    row = 4
    ws4[f'A{row}'] = 'Metric'
    ws4[f'B{row}'] = 'Count'
    ws4[f'A{row}'].font = Font(bold=True)
    ws4[f'B{row}'].font = Font(bold=True)

    stats = [
        ('Total MQ Managers', len(all_mqmanagers)),
        ('Total Gateways', len(all_gateways)),
        ('Internal Gateways', len([g for g in all_gateways if g['Scope'] == 'Internal'])),
        ('External Gateways', len([g for g in all_gateways if g['Scope'] == 'External'])),
        ('Total Connections', len(all_connections)),
        ('Total QLocal Queues', sum(m['QLocal Count'] for m in all_mqmanagers)),
        ('Total QRemote Queues', sum(m['QRemote Count'] for m in all_mqmanagers)),
        ('Total QAlias Queues', sum(m['QAlias Count'] for m in all_mqmanagers))
    ]

    for metric, count in stats:
        row += 1
        ws4[f'A{row}'] = metric
        ws4[f'B{row}'] = count

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 15

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel inventory generated: {output_file}")
    return True
